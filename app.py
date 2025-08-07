from flask import Flask, render_template, request, g, flash, redirect, url_for, session, jsonify, send_file
import sqlite3
from datetime import datetime, timedelta

def to_date_filter(value):
    if not value:
        return None
    return datetime.strptime(value, '%Y-%m-%d').date()

import os
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

app = Flask(__name__)
print(f" * Flask App Root Path: {app.root_path}")
print(f" * Flask App Template Folder: {app.template_folder}")
app.jinja_env.filters['to_date'] = to_date_filter
app.jinja_env.globals.update(now=datetime.now, timedelta=timedelta)
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['DATABASE'] = 'production_schedule.db'
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your_fallback_secret_key_here')

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(app.config['DATABASE'])
        db.row_factory = sqlite3.Row
    return db

def init_db():
    with app.app_context():
        db = get_db()
        cursor = db.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='production_schedule'")
        if cursor.fetchone() is None:
            print("資料庫資料表不存在，正在從 schema.sql 建立...")
            with app.open_resource('schema.sql', mode='r', encoding='utf-8') as f:
                cursor.executescript(f.read())
            db.commit()
            print("資料表建立完成。")

        admin_user = db.execute("SELECT * FROM users WHERE username = 'admin'").fetchone()
        if not admin_user:
            print("正在建立預設 admin 帳號...")
            hashed_password = generate_password_hash('admin')
            db.execute(
                "INSERT INTO users (username, password, name, is_admin) VALUES (?, ?, ?, ?)",
                ('admin', hashed_password, '管理員', 1)
            )
            db.commit()
            print("已創建 admin 帳號，密碼為 'admin'")

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    db = get_db()
    jobs = db.execute("SELECT ps.*, u.name AS creator_name FROM production_schedule ps LEFT JOIN users u ON ps.created_by_user_id = u.id ORDER BY ps.creation_date DESC").fetchall()
    return render_template('index.html', jobs=jobs, current_user_name=session.get('name'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('index'))

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        db = get_db()
        user = db.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()

        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['name'] = user['name']
            session['is_admin'] = user['is_admin']

            ip_address = request.remote_addr
            login_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            db.execute(
                "INSERT INTO login_records (user_id, username, ip_address, login_time) VALUES (?, ?, ?, ?)",
                (user['id'], user['username'], ip_address, login_time)
            )
            db.commit()
            return redirect(url_for('index'))
        else:
            flash("帳號或密碼錯誤！")
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        name = request.form['name']
        hashed_password = generate_password_hash(password)
        db = get_db()
        try:
            db.execute(
                "INSERT INTO users (username, password, name) VALUES (?, ?, ?)",
                (username, hashed_password, name)
            )
            db.commit()
            flash("註冊成功，請登入！")
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            flash("使用者名稱已存在！")
    return render_template('register.html')

@app.route('/logout')
def logout():
    session.clear()
    flash("您已登出。")
    return redirect(url_for('login'))

@app.route('/create_job', methods=['POST'])
def create_job():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    work_order = request.form['work_order']
    model_name = request.form['model_name']
    part_name = request.form['part_name']
    priority = request.form['priority']
    creation_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    created_by_user_id = session.get('user_id')

    db = get_db()
    db.execute(
        "INSERT INTO production_schedule (work_order, model_name, part_name, priority, creation_date, created_by_user_id) VALUES (?, ?, ?, ?, ?, ?)",
        (work_order, model_name, part_name, priority, creation_date, created_by_user_id)
    )
    db.commit()
    flash('工單已成功建立！')
    return redirect(url_for('index'))

@app.route('/edit_job/<int:job_id>', methods=['POST'])
def edit_job(job_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    work_order = request.form['work_order']
    model_name = request.form['model_name']
    part_name = request.form['part_name']
    priority = request.form['priority']

    db = get_db()
    job = db.execute("SELECT * FROM production_schedule WHERE id = ?", (job_id,)).fetchone()

    if job and job['request_date']:
        request_date = datetime.strptime(job['request_date'], '%Y-%m-%d').date()
        if request_date <= (datetime.now() + timedelta(days=1)).date():
            priority = '高'

    db.execute(
        "UPDATE production_schedule SET work_order = ?, model_name = ?, part_name = ?, priority = ? WHERE id = ?",
        (work_order, model_name, part_name, priority, job_id)
    )
    db.commit()
    flash('工單已成功更新！')
    return redirect(url_for('index'))

@app.route('/schedule_job/<int:job_id>', methods=['POST'])
def schedule_job(job_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    material_arrival_date = request.form['material_arrival_date']
    request_date_str = request.form['request_date']
    painting_date = request.form['painting_date']
    notes = request.form['notes']

    priority = '中'  # Default priority
    if request_date_str:
        request_date = datetime.strptime(request_date_str, '%Y-%m-%d').date()
        if request_date <= (datetime.now() + timedelta(days=1)).date():
            priority = '高'

    db = get_db()
    db.execute(
        "UPDATE production_schedule SET material_arrival_date = ?, request_date = ?, painting_date = ?, notes = ?, status = '已排程', priority = ? WHERE id = ?",
        (material_arrival_date, request_date_str, painting_date, notes, priority, job_id)
    )
    db.commit()
    flash('工單已成功排程！')
    return redirect(url_for('index'))

@app.route('/update_status/<int:job_id>', methods=['POST'])
def update_status(job_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    status = request.form['status']
    db = get_db()
    db.execute("UPDATE production_schedule SET status = ? WHERE id = ?", (status, job_id))
    db.commit()
    flash('工單狀態已更新！')
    return redirect(url_for('index'))

@app.route('/delete_job/<int:job_id>', methods=['POST'])
def delete_job(job_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    db = get_db()
    db.execute("DELETE FROM production_schedule WHERE id = ?", (job_id,))
    db.commit()
    flash('工單已成功刪除！')
    return redirect(url_for('index'))

@app.route('/autocomplete_work_orders')
def autocomplete_work_orders():
    if 'user_id' not in session:
        return jsonify([])

    term = request.args.get('term', '')
    db = get_db()
    # 限制查詢結果數量，避免返回過多數據
    work_orders = db.execute("SELECT DISTINCT work_order FROM production_schedule WHERE work_order LIKE ? LIMIT 10", ('%' + term + '%',)).fetchall()
    return jsonify([wo['work_order'] for wo in work_orders])

@app.route('/query_job', methods=['GET', 'POST'])
def query_job():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    jobs = []
    if request.method == 'POST':
        work_order = request.form['work_order']
        db = get_db()
        jobs = db.execute("SELECT ps.*, u.name AS creator_name FROM production_schedule ps LEFT JOIN users u ON ps.created_by_user_id = u.id WHERE ps.work_order LIKE ? ORDER BY ps.creation_date DESC", ('%' + work_order + '%',)).fetchall()
        if not jobs:
            flash(f"找不到工單號碼 '{work_order}' 的排程記錄。")
    return render_template('query_job.html', jobs=jobs, current_user_name=session.get('name'))

@app.route('/admin')
def admin():
    if 'user_id' not in session or session.get('is_admin') != 1:
        flash("您沒有權限訪問此頁面！")
        return redirect(url_for('index'))

    db = get_db()
    login_records = db.execute("SELECT username, ip_address, login_time FROM login_records ORDER BY login_time DESC").fetchall()
    return render_template('admin.html', login_records=login_records, current_user_name=session.get('name'))

@app.route('/manage_users')
def manage_users():
    if 'user_id' not in session or session.get('is_admin') != 1:
        flash("您沒有權限訪問此頁面！")
        return redirect(url_for('index'))
    
    db = get_db()
    users = db.execute("SELECT id, username, name, is_admin FROM users ORDER BY id").fetchall()
    return render_template('manage_users.html', users=users, current_user_name=session.get('name'))

@app.route('/delete_user/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    if 'user_id' not in session or session.get('is_admin') != 1:
        flash("您沒有權限執行此操作！", 'error')
        return redirect(url_for('manage_users'))

    db = get_db()
    user_to_delete = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()

    if user_to_delete is None:
        flash("找不到該使用者！", 'error')
        return redirect(url_for('manage_users'))

    if user_to_delete['username'] == 'admin':
        flash("不能刪除預設的 admin 管理員帳號！", 'error')
        return redirect(url_for('manage_users'))

    # 在刪除使用者之前，處理關聯的資料
    # 例如，將他們建立的工單 created_by_user_id 設為 NULL
    db.execute("UPDATE production_schedule SET created_by_user_id = NULL WHERE created_by_user_id = ?", (user_id,))
    # 刪除使用者的登入記錄
    db.execute("DELETE FROM login_records WHERE user_id = ?", (user_id,))
    # 最後刪除使用者
    db.execute("DELETE FROM users WHERE id = ?", (user_id,))
    db.commit()

    flash(f"使用者 '{user_to_delete['username']}' 已成功刪除。", 'success')
    return redirect(url_for('manage_users'))

@app.route('/export_excel')
def export_excel():
    if 'user_id' not in session or session.get('is_admin') != 1:
        flash("您沒有權限匯出 Excel 檔案！")
        return redirect(url_for('index'))

    db = get_db()
    query = """
    SELECT 
        ps.id,
        ps.work_order,
        ps.model_name,
        ps.part_name,
        ps.customer,
        ps.creation_date,
        ps.material_arrival_date,
        ps.request_date,
        ps.painting_date,
        ps.status,
        ps.priority,
        ps.notes,
        u.name AS creator_name
    FROM 
        production_schedule ps
    LEFT JOIN 
        users u ON ps.created_by_user_id = u.id
    """
    df = pd.read_sql_query(query, db)

    # Translate column headers to Chinese
    column_mapping = {
        'id': 'ID',
        'work_order': '工單號碼',
        'model_name': '機型',
        'part_name': '零件名稱',
        'customer': '客戶',
        'creation_date': '建立日期',
        'material_arrival_date': '入料日期',
        'request_date': '需求日期',
        'painting_date': '噴漆日期',
        'status': '狀態',
        'priority': '優先級',
        'notes': '備註',
        'creator_name': '建立者'
    }
    df.rename(columns=column_mapping, inplace=True)

    # Create an in-memory Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']
        for i, col in enumerate(df.columns):
            column_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
            worksheet.column_dimensions[get_column_letter(i + 1)].width = column_len
    output.seek(0)

    return send_file(output, as_attachment=True, download_name='production_schedule.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')