
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DB_PATH = 'production_schedule.db'
EXCEL_PATH = 'production_schedule.xlsx'

def export_to_excel():
    conn = sqlite3.connect(DB_PATH)
    
    # Read data into a pandas DataFrame
    df = pd.read_sql_query("SELECT * FROM production_schedule", conn)
    
    # Write DataFrame to an Excel file
    with pd.ExcelWriter(EXCEL_PATH, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

        # Auto-adjust column widths
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    conn.close()
    print(f"Data exported to {EXCEL_PATH}")

if __name__ == '__main__':
    export_to_excel()
